<<<<<<< HEAD
"""
Document Intelligence Pipeline
================================
Reads file names from an Excel file (first column),
classifies each document using OpenAI GPT-4o-mini,
and writes results back to a new Excel output file.

Requirements:
    pip install openai pdfplumber openpyxl pandas pytesseract pillow

Usage:
    python classify_documents.py --input files.xlsx
"""

import os
import sys
import json
import time
import argparse
import logging
from pathlib import Path

import pandas as pd
import pdfplumber
import openpyxl
from openai import OpenAI

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
OPENAI_MODEL = "gpt-4o-mini"

# Tier 1 document labels
DOC_LABELS = [
    "Rent Roll",
    "Trial Balance",
    "Budget",
    "Bank Reconciliation",
    "PM Package",
    "Other",
]

# Tier 2 section labels (used only when doc_type == "PM Package")
SECTION_LABELS = [
    "General Ledger",
    "AR Ledger",
    "AP Ledger",
    "Bank Reconciliation",
    "Rent Roll",
    "Vacancy Report",
    "Owner Statement",
    "Cover Page",
    "Other",
]

# Confidence threshold — files below this go to review queue
CONFIDENCE_THRESHOLD = 0.88

# Max lines extracted per page for classification
LINES_PER_PAGE = 5

# ── OpenAI client ─────────────────────────────────────────────────────────────
client = OpenAI()   # reads OPENAI_API_KEY from environment automatically


# ─────────────────────────────────────────────────────────────────────────────
# TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_pdf(filepath: str) -> dict:
    """
    Extract the first LINES_PER_PAGE lines from each page of a PDF.
    Falls back to OCR if no text layer is detected on a page.
    Returns:
        {
          "pages": [{"page": 1, "text": "..."}, ...],
          "used_ocr": True/False
        }
    """
    pages = []
    used_ocr = False

    try:
        with pdfplumber.open(filepath) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                lines = [l.strip() for l in text.split("\n") if l.strip()]

                # If no text layer, try OCR
                if not lines:
                    try:
                        import pytesseract
                        from PIL import Image
                        img = page.to_image(resolution=150).original
                        text = pytesseract.image_to_string(img)
                        lines = [l.strip() for l in text.split("\n") if l.strip()]
                        used_ocr = True
                        log.debug(f"  OCR used on page {i+1} of {filepath}")
                    except ImportError:
                        log.warning("  pytesseract not installed — skipping OCR for this page")
                    except Exception as e:
                        log.warning(f"  OCR failed on page {i+1}: {e}")

                pages.append({
                    "page": i + 1,
                    "text": "\n".join(lines[:LINES_PER_PAGE])
                })

    except Exception as e:
        log.error(f"  PDF extraction failed: {e}")

    return {"pages": pages, "used_ocr": used_ocr}


def extract_excel(filepath: str) -> dict:
    """
    Extract sheet names, column headers, and first 3 data rows from each sheet.
    """
    pages = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(max_row=5, values_only=True):
                row_text = "  |  ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                pages.append({
                    "page": sheet_name,
                    "text": "\n".join(rows)
                })
        wb.close()
    except Exception as e:
        log.error(f"  Excel extraction failed: {e}")
    return {"pages": pages, "used_ocr": False}


def extract_csv(filepath: str) -> dict:
    """
    Extract column headers and first 5 rows from a CSV file.
    """
    pages = []
    try:
        df = pd.read_csv(filepath, nrows=5, dtype=str)
        header = "  |  ".join(df.columns.tolist())
        rows = [header]
        for _, row in df.iterrows():
            rows.append("  |  ".join(row.fillna("").tolist()))
        pages.append({"page": 1, "text": "\n".join(rows)})
    except Exception as e:
        log.error(f"  CSV extraction failed: {e}")
    return {"pages": pages, "used_ocr": False}


def extract_file(filepath: str) -> dict:
    """Route to the correct extractor based on file extension."""
    ext = Path(filepath).suffix.lower()
    if ext == ".pdf":
        return extract_pdf(filepath)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        return extract_excel(filepath)
    elif ext == ".csv":
        return extract_csv(filepath)
    else:
        log.warning(f"  Unsupported file type: {ext}")
        return {"pages": [], "used_ocr": False}


def build_extract_text(extraction: dict) -> str:
    """Format extracted pages into a single string for the LLM prompt."""
    parts = []
    for p in extraction["pages"]:
        parts.append(f"--- Page/Sheet: {p['page']} ---\n{p['text']}")
    return "\n\n".join(parts) if parts else "(No readable text found)"


# ─────────────────────────────────────────────────────────────────────────────
# TIER 1 — DOCUMENT CLASSIFICATION
# ─────────────────────────────────────────────────────────────────────────────

TIER1_SYSTEM = """You are a financial document classifier for a real estate property management company.
You receive the first few lines from each page of a document and must classify the overall document type.
Always respond with valid JSON only — no explanation, no markdown, no code fences."""

TIER1_USER_TEMPLATE = """Classify this document and return JSON in exactly this format:
{{
  "label": "<one of: {labels}>",
  "confidence": <float 0.0 to 1.0>,
  "reason": "<one sentence explaining your confidence level>",
  "property": "<property name if identifiable, else null>",
  "period": "<reporting period if identifiable e.g. Feb 2026, else null>"
}}

Document extract:
{text}"""


def classify_document(extract_text: str, retries: int = 3) -> dict:
    """
    Tier 1: classify the whole document.
    Returns a dict with label, confidence, reason, property, period.
    """
    prompt = TIER1_USER_TEMPLATE.format(
        labels=", ".join(DOC_LABELS),
        text=extract_text[:6000]   # hard cap — never send huge text
    )

    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": TIER1_SYSTEM},
                    {"role": "user",   "content": prompt},
                ],
                temperature=0,
                max_tokens=300,
            )
            raw = response.choices[0].message.content.strip()
            result = json.loads(raw)
            # Validate required keys
            for key in ("label", "confidence", "reason"):
                if key not in result:
                    raise ValueError(f"Missing key: {key}")
            return result

        except (json.JSONDecodeError, ValueError) as e:
            log.warning(f"  Tier1 parse error (attempt {attempt+1}): {e}")
            time.sleep(1)
        except Exception as e:
            log.warning(f"  Tier1 API error (attempt {attempt+1}): {e}")
            time.sleep(2 ** attempt)

    return {
        "label": "Other",
        "confidence": 0.0,
        "reason": "Classification failed after retries.",
        "property": None,
        "period": None,
    }


# ─────────────────────────────────────────────────────────────────────────────
# TIER 2 — SECTION CLASSIFICATION (PM PACKAGES ONLY)
# ─────────────────────────────────────────────────────────────────────────────

TIER2_SYSTEM = """You are a financial document section classifier for a real estate property management company.
You receive the first few lines from each page of a PM Package and must classify what section each page belongs to.
Always respond with valid JSON only — no explanation, no markdown, no code fences."""

TIER2_USER_TEMPLATE = """A PM Package has {num_pages} pages. For each page below, classify which financial section it belongs to.

Valid section labels: {labels}

Return a JSON array with one object per page:
[
  {{"page": 1, "section": "<label>", "confidence": <0.0-1.0>}},
  ...
]

Page extracts:
{text}"""


def classify_sections(pages: list, retries: int = 3) -> list:
    """
    Tier 2: classify each page of a PM Package.
    Returns list of dicts: [{"page": N, "section": "...", "confidence": ...}]
    """
    # Build condensed text — one entry per page
    text_parts = []
    for p in pages:
        text_parts.append(f"[Page {p['page']}]\n{p['text'][:400]}")
    full_text = "\n\n".join(text_parts)

    prompt = TIER2_USER_TEMPLATE.format(
        num_pages=len(pages),
        labels=", ".join(SECTION_LABELS),
        text=full_text[:8000]
    )

    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": TIER2_SYSTEM},
                    {"role": "user",   "content": prompt},
                ],
                temperature=0,
                max_tokens=1000,
            )
            raw = response.choices[0].message.content.strip()
            result = json.loads(raw)
            if isinstance(result, list):
                return result
            raise ValueError("Expected a JSON array")

        except (json.JSONDecodeError, ValueError) as e:
            log.warning(f"  Tier2 parse error (attempt {attempt+1}): {e}")
            time.sleep(1)
        except Exception as e:
            log.warning(f"  Tier2 API error (attempt {attempt+1}): {e}")
            time.sleep(2 ** attempt)

    # Fallback: mark every page as Other
    return [{"page": p["page"], "section": "Other", "confidence": 0.0} for p in pages]


def merge_sections(page_labels: list) -> list:
    """
    Collapse consecutive pages with the same section label into a single
    section record with a start_page and end_page range.

    Input:  [{"page":1,"section":"GL",...}, {"page":2,"section":"GL",...}, ...]
    Output: [{"section":"GL","start_page":1,"end_page":2,"avg_confidence":0.95}, ...]
    """
    if not page_labels:
        return []

    merged = []
    current = dict(page_labels[0])
    current["start_page"] = current.pop("page")
    current["end_page"]   = current["start_page"]
    confidences = [current.get("confidence", 0)]

    for entry in page_labels[1:]:
        if entry["section"] == current["section"]:
            current["end_page"] = entry["page"]
            confidences.append(entry.get("confidence", 0))
        else:
            current["avg_confidence"] = round(sum(confidences) / len(confidences), 3)
            current.pop("confidence", None)
            merged.append(current)
            current = dict(entry)
            current["start_page"] = current.pop("page")
            current["end_page"]   = current["start_page"]
            confidences = [current.get("confidence", 0)]

    current["avg_confidence"] = round(sum(confidences) / len(confidences), 3)
    current.pop("confidence", None)
    merged.append(current)
    return merged


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_file(filename: str, base_dir: str) -> dict:
    """
    Full pipeline for a single file.
    Returns a result dict ready to be written to the output Excel.
    """
    filepath = os.path.join(base_dir, filename)

    result = {
        "filename":        filename,
        "doc_type":        None,
        "property":        None,
        "period":          None,
        "confidence":      None,
        "status":          None,
        "reason":          None,
        "used_ocr":        False,
        "sections":        None,   # JSON string — only for PM Packages
        "error":           None,
    }

    # ── File existence check ──────────────────────────────────────────────────
    if not os.path.isfile(filepath):
        result["error"]  = "File not found"
        result["status"] = "error"
        log.warning(f"  File not found: {filepath}")
        return result

    log.info(f"  Extracting  →  {filename}")

    # ── Stage 1: Extract ──────────────────────────────────────────────────────
    extraction = extract_file(filepath)
    result["used_ocr"] = extraction.get("used_ocr", False)

    if not extraction["pages"]:
        result["error"]  = "No text could be extracted"
        result["status"] = "error"
        return result

    extract_text = build_extract_text(extraction)

    # ── Stage 2: Tier 1 classification ───────────────────────────────────────
    log.info(f"  Classifying →  {filename}")
    tier1 = classify_document(extract_text)

    result["doc_type"]   = tier1.get("label",      "Other")
    result["confidence"] = tier1.get("confidence", 0.0)
    result["reason"]     = tier1.get("reason",     "")
    result["property"]   = tier1.get("property")
    result["period"]     = tier1.get("period")
    result["status"]     = (
        "auto-labeled"    if result["confidence"] >= CONFIDENCE_THRESHOLD
        else "needs-review" if result["confidence"] >= 0.65
        else "unclassifiable"
    )

    # ── Stage 3: Tier 2 section classification (PM Packages only) ────────────
    if result["doc_type"] == "PM Package":
        log.info(f"  PM Package detected — running section classifier on {len(extraction['pages'])} pages")
        page_labels = classify_sections(extraction["pages"])
        sections    = merge_sections(page_labels)
        result["sections"] = json.dumps(sections, indent=None)
        log.info(f"  Sections found: {[s['section'] for s in sections]}")

    return result


def run_pipeline(input_excel: str):
    """
    Read filenames from the first column of input_excel,
    process each file, and write results to a new Excel file.
    """
    base_dir   = os.path.dirname(os.path.abspath(input_excel))
    input_path = os.path.abspath(input_excel)

    # ── Read input ────────────────────────────────────────────────────────────
    log.info(f"Reading input file: {input_path}")
    try:
        df_input = pd.read_excel(input_path, dtype=str)
    except Exception as e:
        log.error(f"Cannot read input Excel: {e}")
        sys.exit(1)

    first_col = df_input.columns[0]
    filenames = df_input[first_col].dropna().str.strip().tolist()
    log.info(f"Found {len(filenames)} file(s) to process")

    # ── Process each file ─────────────────────────────────────────────────────
    results = []
    for idx, filename in enumerate(filenames, 1):
        log.info(f"[{idx}/{len(filenames)}] {filename}")
        result = process_file(filename, base_dir)
        results.append(result)
        time.sleep(0.3)  # gentle rate-limit buffer

    # ── Build output DataFrame ────────────────────────────────────────────────
    df_out = pd.DataFrame(results, columns=[
        "filename", "doc_type", "property", "period",
        "confidence", "status", "reason", "used_ocr", "sections", "error"
    ])

    # Rename columns to friendlier display names
    df_out.columns = [
        "File Name", "Document Type", "Property", "Period",
        "Confidence", "Status", "Reason", "OCR Used", "Sections (PM Package)", "Error"
    ]

    # ── Write output Excel ────────────────────────────────────────────────────
    output_path = os.path.join(
        base_dir,
        Path(input_excel).stem + "_classified.xlsx"
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Classification Results")

        ws = writer.sheets["Classification Results"]

        # Style the header row
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            bottom=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin",  color="CCCCCC"),
        )

        status_colors = {
            "auto-labeled":     "E2EFDA",  # light green
            "needs-review":     "FFF2CC",  # light amber
            "unclassifiable":   "FCE4D6",  # light red/orange
            "error":            "F4CCCC",  # light red
        }

        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Style data rows and apply status colours
        status_col_idx = 6   # "Status" is the 6th column (1-indexed)
        for row in ws.iter_rows(min_row=2):
            status_val = str(row[status_col_idx - 1].value or "").lower()
            row_fill_color = status_colors.get(status_val)
            for cell in row:
                cell.font      = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border    = thin_border
                if row_fill_color:
                    cell.fill = PatternFill(start_color=row_fill_color,
                                            end_color=row_fill_color,
                                            fill_type="solid")

        # Auto-fit column widths (approximate)
        col_widths = [40, 20, 25, 15, 12, 16, 50, 10, 60, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Freeze top row
        ws.freeze_panes = "A2"

    log.info(f"\nDone. Output written to: {output_path}")

    # ── Summary ───────────────────────────────────────────────────────────────
    total       = len(df_out)
    auto        = (df_out["Status"] == "auto-labeled").sum()
    review      = (df_out["Status"] == "needs-review").sum()
    unclear     = (df_out["Status"] == "unclassifiable").sum()
    errors      = (df_out["Status"] == "error").sum()
    pm_packages = (df_out["Document Type"] == "PM Package").sum()

    print("\n" + "="*50)
    print("  CLASSIFICATION SUMMARY")
    print("="*50)
    print(f"  Total files processed : {total}")
    print(f"  Auto-labeled          : {auto}")
    print(f"  Needs review          : {review}")
    print(f"  Unclassifiable        : {unclear}")
    print(f"  Errors                : {errors}")
    print(f"  PM Packages found     : {pm_packages}")
    print(f"  Output file           : {output_path}")
    print("="*50 + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Classify financial documents using GPT-4o-mini."
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Path to the Excel file containing file names in the first column."
    )
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"Error: input file not found: {args.input}")
        sys.exit(1)

=======
"""
Document Intelligence Pipeline
================================
Reads file names from an Excel file (first column),
classifies each document using OpenAI GPT-4o-mini,
and writes results back to a new Excel output file.

Requirements:
    pip install openai pdfplumber openpyxl pandas pytesseract pillow

Usage:
    python classify_documents.py --input files.xlsx
"""

import os
import sys
import json
import time
import argparse
import logging
from pathlib import Path

import pandas as pd
import pdfplumber
import openpyxl
from openai import OpenAI

# ── Logging ──────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
OPENAI_MODEL = "gpt-4o-mini"

# Tier 1 document labels
DOC_LABELS = [
    "Rent Roll",
    "Trial Balance",
    "Budget",
    "Bank Reconciliation",
    "PM Package",
    "Other",
]

# Tier 2 section labels (used only when doc_type == "PM Package")
SECTION_LABELS = [
    "General Ledger",
    "AR Ledger",
    "AP Ledger",
    "Bank Reconciliation",
    "Rent Roll",
    "Vacancy Report",
    "Owner Statement",
    "Cover Page",
    "Other",
]

# Confidence threshold — files below this go to review queue
CONFIDENCE_THRESHOLD = 0.88

# Max lines extracted per page for classification
LINES_PER_PAGE = 5

# ── OpenAI client ─────────────────────────────────────────────────────────────
client = OpenAI()   # reads OPENAI_API_KEY from environment automatically


# ─────────────────────────────────────────────────────────────────────────────
# TEXT EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def extract_pdf(filepath: str) -> dict:
    """
    Extract the first LINES_PER_PAGE lines from each page of a PDF.
    Falls back to OCR if no text layer is detected on a page.
    Returns:
        {
          "pages": [{"page": 1, "text": "..."}, ...],
          "used_ocr": True/False
        }
    """
    pages = []
    used_ocr = False

    try:
        with pdfplumber.open(filepath) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                lines = [l.strip() for l in text.split("\n") if l.strip()]

                # If no text layer, try OCR
                if not lines:
                    try:
                        import pytesseract
                        from PIL import Image
                        img = page.to_image(resolution=150).original
                        text = pytesseract.image_to_string(img)
                        lines = [l.strip() for l in text.split("\n") if l.strip()]
                        used_ocr = True
                        log.debug(f"  OCR used on page {i+1} of {filepath}")
                    except ImportError:
                        log.warning("  pytesseract not installed — skipping OCR for this page")
                    except Exception as e:
                        log.warning(f"  OCR failed on page {i+1}: {e}")

                pages.append({
                    "page": i + 1,
                    "text": "\n".join(lines[:LINES_PER_PAGE])
                })

    except Exception as e:
        log.error(f"  PDF extraction failed: {e}")

    return {"pages": pages, "used_ocr": used_ocr}


def extract_excel(filepath: str) -> dict:
    """
    Extract sheet names, column headers, and first 3 data rows from each sheet.
    """
    pages = []
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(max_row=5, values_only=True):
                row_text = "  |  ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                pages.append({
                    "page": sheet_name,
                    "text": "\n".join(rows)
                })
        wb.close()
    except Exception as e:
        log.error(f"  Excel extraction failed: {e}")
    return {"pages": pages, "used_ocr": False}


def extract_csv(filepath: str) -> dict:
    """
    Extract column headers and first 5 rows from a CSV file.
    """
    pages = []
    try:
        df = pd.read_csv(filepath, nrows=5, dtype=str)
        header = "  |  ".join(df.columns.tolist())
        rows = [header]
        for _, row in df.iterrows():
            rows.append("  |  ".join(row.fillna("").tolist()))
        pages.append({"page": 1, "text": "\n".join(rows)})
    except Exception as e:
        log.error(f"  CSV extraction failed: {e}")
    return {"pages": pages, "used_ocr": False}


def extract_file(filepath: str) -> dict:
    """Route to the correct extractor based on file extension."""
    ext = Path(filepath).suffix.lower()
    if ext == ".pdf":
        return extract_pdf(filepath)
    elif ext in (".xlsx", ".xlsm", ".xls"):
        return extract_excel(filepath)
    elif ext == ".csv":
        return extract_csv(filepath)
    else:
        log.warning(f"  Unsupported file type: {ext}")
        return {"pages": [], "used_ocr": False}


def build_extract_text(extraction: dict) -> str:
    """Format extracted pages into a single string for the LLM prompt."""
    parts = []
    for p in extraction["pages"]:
        parts.append(f"--- Page/Sheet: {p['page']} ---\n{p['text']}")
    return "\n\n".join(parts) if parts else "(No readable text found)"


# ─────────────────────────────────────────────────────────────────────────────
# TIER 1 — DOCUMENT CLASSIFICATION
# ─────────────────────────────────────────────────────────────────────────────

TIER1_SYSTEM = """You are a financial document classifier for a real estate property management company.
You receive the first few lines from each page of a document and must classify the overall document type.
Always respond with valid JSON only — no explanation, no markdown, no code fences."""

TIER1_USER_TEMPLATE = """Classify this document and return JSON in exactly this format:
{{
  "label": "<one of: {labels}>",
  "confidence": <float 0.0 to 1.0>,
  "reason": "<one sentence explaining your confidence level>",
  "property": "<property name if identifiable, else null>",
  "period": "<reporting period if identifiable e.g. Feb 2026, else null>"
}}

Document extract:
{text}"""


def classify_document(extract_text: str, retries: int = 3) -> dict:
    """
    Tier 1: classify the whole document.
    Returns a dict with label, confidence, reason, property, period.
    """
    prompt = TIER1_USER_TEMPLATE.format(
        labels=", ".join(DOC_LABELS),
        text=extract_text[:6000]   # hard cap — never send huge text
    )

    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": TIER1_SYSTEM},
                    {"role": "user",   "content": prompt},
                ],
                temperature=0,
                max_tokens=300,
            )
            raw = response.choices[0].message.content.strip()
            result = json.loads(raw)
            # Validate required keys
            for key in ("label", "confidence", "reason"):
                if key not in result:
                    raise ValueError(f"Missing key: {key}")
            return result

        except (json.JSONDecodeError, ValueError) as e:
            log.warning(f"  Tier1 parse error (attempt {attempt+1}): {e}")
            time.sleep(1)
        except Exception as e:
            log.warning(f"  Tier1 API error (attempt {attempt+1}): {e}")
            time.sleep(2 ** attempt)

    return {
        "label": "Other",
        "confidence": 0.0,
        "reason": "Classification failed after retries.",
        "property": None,
        "period": None,
    }


# ─────────────────────────────────────────────────────────────────────────────
# TIER 2 — SECTION CLASSIFICATION (PM PACKAGES ONLY)
# ─────────────────────────────────────────────────────────────────────────────

TIER2_SYSTEM = """You are a financial document section classifier for a real estate property management company.
You receive the first few lines from each page of a PM Package and must classify what section each page belongs to.
Always respond with valid JSON only — no explanation, no markdown, no code fences."""

TIER2_USER_TEMPLATE = """A PM Package has {num_pages} pages. For each page below, classify which financial section it belongs to.

Valid section labels: {labels}

Return a JSON array with one object per page:
[
  {{"page": 1, "section": "<label>", "confidence": <0.0-1.0>}},
  ...
]

Page extracts:
{text}"""


def classify_sections(pages: list, retries: int = 3) -> list:
    """
    Tier 2: classify each page of a PM Package.
    Returns list of dicts: [{"page": N, "section": "...", "confidence": ...}]
    """
    # Build condensed text — one entry per page
    text_parts = []
    for p in pages:
        text_parts.append(f"[Page {p['page']}]\n{p['text'][:400]}")
    full_text = "\n\n".join(text_parts)

    prompt = TIER2_USER_TEMPLATE.format(
        num_pages=len(pages),
        labels=", ".join(SECTION_LABELS),
        text=full_text[:8000]
    )

    for attempt in range(retries):
        try:
            response = client.chat.completions.create(
                model=OPENAI_MODEL,
                messages=[
                    {"role": "system", "content": TIER2_SYSTEM},
                    {"role": "user",   "content": prompt},
                ],
                temperature=0,
                max_tokens=1000,
            )
            raw = response.choices[0].message.content.strip()
            result = json.loads(raw)
            if isinstance(result, list):
                return result
            raise ValueError("Expected a JSON array")

        except (json.JSONDecodeError, ValueError) as e:
            log.warning(f"  Tier2 parse error (attempt {attempt+1}): {e}")
            time.sleep(1)
        except Exception as e:
            log.warning(f"  Tier2 API error (attempt {attempt+1}): {e}")
            time.sleep(2 ** attempt)

    # Fallback: mark every page as Other
    return [{"page": p["page"], "section": "Other", "confidence": 0.0} for p in pages]


def merge_sections(page_labels: list) -> list:
    """
    Collapse consecutive pages with the same section label into a single
    section record with a start_page and end_page range.

    Input:  [{"page":1,"section":"GL",...}, {"page":2,"section":"GL",...}, ...]
    Output: [{"section":"GL","start_page":1,"end_page":2,"avg_confidence":0.95}, ...]
    """
    if not page_labels:
        return []

    merged = []
    current = dict(page_labels[0])
    current["start_page"] = current.pop("page")
    current["end_page"]   = current["start_page"]
    confidences = [current.get("confidence", 0)]

    for entry in page_labels[1:]:
        if entry["section"] == current["section"]:
            current["end_page"] = entry["page"]
            confidences.append(entry.get("confidence", 0))
        else:
            current["avg_confidence"] = round(sum(confidences) / len(confidences), 3)
            current.pop("confidence", None)
            merged.append(current)
            current = dict(entry)
            current["start_page"] = current.pop("page")
            current["end_page"]   = current["start_page"]
            confidences = [current.get("confidence", 0)]

    current["avg_confidence"] = round(sum(confidences) / len(confidences), 3)
    current.pop("confidence", None)
    merged.append(current)
    return merged


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def process_file(filename: str, base_dir: str) -> dict:
    """
    Full pipeline for a single file.
    Returns a result dict ready to be written to the output Excel.
    """
    filepath = os.path.join(base_dir, filename)

    result = {
        "filename":        filename,
        "doc_type":        None,
        "property":        None,
        "period":          None,
        "confidence":      None,
        "status":          None,
        "reason":          None,
        "used_ocr":        False,
        "sections":        None,   # JSON string — only for PM Packages
        "error":           None,
    }

    # ── File existence check ──────────────────────────────────────────────────
    if not os.path.isfile(filepath):
        result["error"]  = "File not found"
        result["status"] = "error"
        log.warning(f"  File not found: {filepath}")
        return result

    log.info(f"  Extracting  →  {filename}")

    # ── Stage 1: Extract ──────────────────────────────────────────────────────
    extraction = extract_file(filepath)
    result["used_ocr"] = extraction.get("used_ocr", False)

    if not extraction["pages"]:
        result["error"]  = "No text could be extracted"
        result["status"] = "error"
        return result

    extract_text = build_extract_text(extraction)

    # ── Stage 2: Tier 1 classification ───────────────────────────────────────
    log.info(f"  Classifying →  {filename}")
    tier1 = classify_document(extract_text)

    result["doc_type"]   = tier1.get("label",      "Other")
    result["confidence"] = tier1.get("confidence", 0.0)
    result["reason"]     = tier1.get("reason",     "")
    result["property"]   = tier1.get("property")
    result["period"]     = tier1.get("period")
    result["status"]     = (
        "auto-labeled"    if result["confidence"] >= CONFIDENCE_THRESHOLD
        else "needs-review" if result["confidence"] >= 0.65
        else "unclassifiable"
    )

    # ── Stage 3: Tier 2 section classification (PM Packages only) ────────────
    if result["doc_type"] == "PM Package":
        log.info(f"  PM Package detected — running section classifier on {len(extraction['pages'])} pages")
        page_labels = classify_sections(extraction["pages"])
        sections    = merge_sections(page_labels)
        result["sections"] = json.dumps(sections, indent=None)
        log.info(f"  Sections found: {[s['section'] for s in sections]}")

    return result


def run_pipeline(input_excel: str):
    """
    Read filenames from the first column of input_excel,
    process each file, and write results to a new Excel file.
    """
    base_dir   = os.path.dirname(os.path.abspath(input_excel))
    input_path = os.path.abspath(input_excel)

    # ── Read input ────────────────────────────────────────────────────────────
    log.info(f"Reading input file: {input_path}")
    try:
        df_input = pd.read_excel(input_path, dtype=str)
    except Exception as e:
        log.error(f"Cannot read input Excel: {e}")
        sys.exit(1)

    first_col = df_input.columns[0]
    filenames = df_input[first_col].dropna().str.strip().tolist()
    log.info(f"Found {len(filenames)} file(s) to process")

    # ── Process each file ─────────────────────────────────────────────────────
    results = []
    for idx, filename in enumerate(filenames, 1):
        log.info(f"[{idx}/{len(filenames)}] {filename}")
        result = process_file(filename, base_dir)
        results.append(result)
        time.sleep(0.3)  # gentle rate-limit buffer

    # ── Build output DataFrame ────────────────────────────────────────────────
    df_out = pd.DataFrame(results, columns=[
        "filename", "doc_type", "property", "period",
        "confidence", "status", "reason", "used_ocr", "sections", "error"
    ])

    # Rename columns to friendlier display names
    df_out.columns = [
        "File Name", "Document Type", "Property", "Period",
        "Confidence", "Status", "Reason", "OCR Used", "Sections (PM Package)", "Error"
    ]

    # ── Write output Excel ────────────────────────────────────────────────────
    output_path = os.path.join(
        base_dir,
        Path(input_excel).stem + "_classified.xlsx"
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Classification Results")

        ws = writer.sheets["Classification Results"]

        # Style the header row
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            bottom=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin",  color="CCCCCC"),
        )

        status_colors = {
            "auto-labeled":     "E2EFDA",  # light green
            "needs-review":     "FFF2CC",  # light amber
            "unclassifiable":   "FCE4D6",  # light red/orange
            "error":            "F4CCCC",  # light red
        }

        for col_idx, cell in enumerate(ws[1], 1):
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Style data rows and apply status colours
        status_col_idx = 6   # "Status" is the 6th column (1-indexed)
        for row in ws.iter_rows(min_row=2):
            status_val = str(row[status_col_idx - 1].value or "").lower()
            row_fill_color = status_colors.get(status_val)
            for cell in row:
                cell.font      = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border    = thin_border
                if row_fill_color:
                    cell.fill = PatternFill(start_color=row_fill_color,
                                            end_color=row_fill_color,
                                            fill_type="solid")

        # Auto-fit column widths (approximate)
        col_widths = [40, 20, 25, 15, 12, 16, 50, 10, 60, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Freeze top row
        ws.freeze_panes = "A2"

    log.info(f"\nDone. Output written to: {output_path}")

    # ── Summary ───────────────────────────────────────────────────────────────
    total       = len(df_out)
    auto        = (df_out["Status"] == "auto-labeled").sum()
    review      = (df_out["Status"] == "needs-review").sum()
    unclear     = (df_out["Status"] == "unclassifiable").sum()
    errors      = (df_out["Status"] == "error").sum()
    pm_packages = (df_out["Document Type"] == "PM Package").sum()

    print("\n" + "="*50)
    print("  CLASSIFICATION SUMMARY")
    print("="*50)
    print(f"  Total files processed : {total}")
    print(f"  Auto-labeled          : {auto}")
    print(f"  Needs review          : {review}")
    print(f"  Unclassifiable        : {unclear}")
    print(f"  Errors                : {errors}")
    print(f"  PM Packages found     : {pm_packages}")
    print(f"  Output file           : {output_path}")
    print("="*50 + "\n")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Classify financial documents using GPT-4o-mini."
    )
    parser.add_argument(
        "--input", "-i",
        required=True,
        help="Path to the Excel file containing file names in the first column."
    )
    args = parser.parse_args()

    if not os.path.isfile(args.input):
        print(f"Error: input file not found: {args.input}")
        sys.exit(1)

>>>>>>> 7963b1677f1072da24f30c272e22a3f9578f7f12
    run_pipeline(args.input)